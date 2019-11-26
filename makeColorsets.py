# Copyright (c) 2019 Jinwook Jeon. All rights reserved.

# Permission is hereby granted, free of charge, to any person obtaining a
# copy of this software and associated documentation files (the "Software"),
# to deal in the Software without restriction, including
# without limitation the rights to use, copy, modify, merge, publish,
# distribute, sublicense, and/or sell copies of the Software, and to
# permit persons to whom the Software is furnished to do so, subject to
# the following conditions:

# The above copyright notice and this permission notice shall be included
# in all copies or substantial portions of the Software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS
# OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY
# CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT,
# TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE
# SOFTWARE OR THE USE OR OTHER DEALINGS IN THE SOFTWARE.

from __future__ import print_function
import os
import io
import json
import enum
import argparse
from openpyxl import load_workbook
from collections import OrderedDict


class IdiomType(enum.Enum):
    UNIVERSAL = 'universal'
    IPHONE = 'iphone'
    IPAD = 'ipad'
    CARPLAY = 'car'
    WATCH = 'watch'
    TV = 'tv'
    MAC = 'mac'


class ColorExtension():
    def __init__(self, name, lightStr, darkStr):
        self.name = name
        self.lightStr = lightStr
        self.darkStr = darkStr


class ColorComponent():
    # Initializer
    # def __init__(self, name, lightColor='#ffffff', lightColorAlpha=1.0, darkColor='#000000', darkColorAlpha=1.0):
    def __init__(self, row):
        self.name = row[0].value
        self.lightColor = row[1].value
        self.lightColorAlpha = row[2].value
        self.darkColor = row[3].value
        self.darkColorAlpha = row[4].value
        self.extension = ColorExtension(self.name,
                                        self._getExtensionStr(),
                                        self._getExtensionStr(isDark=True))

    # Public Methods
    def getLightComponent(self, isHex=True):
        return self._getComponent(isHex=isHex)

    def getDarkComponent(self, isHex=True):
        return self._getComponent(isHex=isHex, isDark=True)


    # Private Methods
    def _getExtensionStr(self, isDark=False):
        color = isDark and self.darkColor or self.lightColor
        alpha = isDark and self.darkColorAlpha or self.lightColorAlpha

        red = self._getFloat(color[1:3])
        green = self._getFloat(color[3:5])
        blue = self._getFloat(color[5:7])
        return f'UIColor(red: {red}, green: {green}, blue: {blue}, alpha: {alpha})'

    def _getComponent(self, isHex=True, isDark=False):
        color = OrderedDict()
        color["color-space"] = "srgb"
        components = OrderedDict()
        targetColor = isDark and self.darkColor or self.lightColor
        alpha = isDark and self.darkColorAlpha or self.lightColorAlpha

        components["red"] = isHex and self._getHex(targetColor[1:3]) or self._getFloat(targetColor[1:3])
        components["green"] = isHex and self._getHex(targetColor[3:5]) or self._getFloat(targetColor[3:5])
        components["blue"] = isHex and self._getHex(targetColor[5:7]) or self._getFloat(targetColor[5:7])
        components["alpha"] = '{0:1.3f}'.format(float(alpha))
        color["components"] = components
        return color

    def _getHex(self, s):
        if len(s) == 0:
            return '0x00'
        return f'0x{s.upper()}'

    def _getFloat(self, s):
        if len(s) == 0:
            return '0.000'
        return '{0:1.3f}'.format(float.fromhex(s) / 255.0)


def getExtensionStr(info):
    extensionStr = f'''
    static var {info.name}: UIColor {{
        if #available(iOS 13, *) {{
            return UIColor {{ (traitCollection: UITraitCollection) -> UIColor in
                if traitCollection.userInterfaceStyle == .dark {{
                    return {info.darkStr}
                }} else {{
                    return {info.lightStr}
                }}
            }}
        }} else {{
            return {info.lightStr}
        }}
    }}'''
    return extensionStr


def getInfo():
    infoDic = OrderedDict()
    infoDic["version"] = 1
    infoDic["author"] = "xcode"
    return infoDic


def getAppearances():
    appearances = []
    appearanceModel = OrderedDict()
    appearanceModel["appearance"] = "luminosity"
    appearanceModel["value"] = "dark"
    appearances.append(appearanceModel)
    return appearances


def getColorInfo(color, idiom=IdiomType.UNIVERSAL, isDarkMode=False, isMacCatalyst=False):
    colorInfo = OrderedDict()
    colorInfo["idiom"] = idiom.value
    if idiom.value == 'ipad' and isMacCatalyst:
        colorInfo["subtype"] = "mac-catalyst"

    if isDarkMode:
        colorInfo["appearances"] = getAppearances()
        colorInfo["color"] = color.getDarkComponent()
    else:
        colorInfo["color"] = color.getLightComponent()

    return colorInfo


def getAllSheets():
    global allColorCount
    global excelSheetName

    workBook = load_workbook(excelFileName, data_only=True)
    sheetNames = excelSheetName == '' and workBook.sheetnames or [excelSheetName]
    sheets = OrderedDict()
    for sheetName in sheetNames:
        workSheet = workBook[sheetName]
        allColors = []
        for row in workSheet.rows:
            if row[0].value is None or row[0].value == 'name':
                continue
            allColors.append(ColorComponent(row))
        sheets[sheetName] = allColors
        allColorCount += len(allColors)
    return sheets


def parse(color, idioms=[IdiomType.UNIVERSAL], needMacCatalyst=False):
    colorList = []
    for idiom in idioms:
        colorList.append(getColorInfo(color, idiom, isDarkMode=False))
        colorList.append(getColorInfo(color, idiom, isDarkMode=True))
        # 'mac-catalyst' is a subType of 'ipad'
        if idiom.value == 'ipad' and needMacCatalyst:
            colorList.append(getColorInfo(
                color, idiom, isDarkMode=False, isMacCatalyst=True))
            colorList.append(getColorInfo(
                color, idiom, isDarkMode=True, isMacCatalyst=True))

    colorSetDic = OrderedDict()
    colorSetDic["info"] = getInfo()
    colorSetDic["colors"] = colorList

    return colorSetDic


def createDir(filePath):
    if not os.path.exists(filePath):
        os.makedirs(filePath)


def saveToFile(data, filePath='./'):
    fileName = f'{filePath}/{jsonFile}'
    with io.open(fileName, 'w+', encoding='utf-8') as file:
        file.write(json.dumps(data, ensure_ascii=False, indent=2, separators=(',', ' : ')))

    print(f'{fileName} file is created')


def makeColorExtension():
    global allSheets

    fileName = 'UIColor+DarkMode.swift'
    with io.open(fileName, 'w+', encoding='utf-8') as file:
        file.write('import UIKit\n\n\n')
        file.write('extension UIColor {')
        for key in allSheets:
            allColors = allSheets[key]
            for color in allColors:
                file.write(getExtensionStr(color.extension))
                file.write('\n')
        file.write('}\n')


def createContainerDir(dirName):
    createDir(dirName)
    colorSetDic = OrderedDict()
    colorSetDic["info"] = getInfo()
    saveToFile(colorSetDic, filePath=dirName)


def printSeparator(count):
    for _ in range(0, count):
        print('=', end="")
    print()


def printHeader(title):
    print()
    printSeparator(len(title) + 1)
    print(title)
    printSeparator(len(title) + 1)
    print()

def extract():
    global allSheets

    allSheets = getAllSheets()
    printHeader(f' {allColorCount} colors are loaded')

    printHeader(' Container folder is creating')
    createContainerDir(rootDir)

    printHeader(' Colorsets are creating')
    for key in allSheets:
        printHeader(f' \'{key}\' folder is creating')
        path = f'{rootDir}/{key}'
        createContainerDir(path)
        allColors = allSheets[key]
        for color in allColors:
            parsed = parse(color, idioms=deviceTypes, needMacCatalyst=macCatalyst)
            path = f'{rootDir}/{key}/{color.name}.colorset'
            createDir(path)
            saveToFile(parsed, filePath=path)
            # print(json.dumps(parsed, ensure_ascii=False, indent="  "))

    makeColorExtension()
    printHeader(f' {allColorCount} colorsets are created')


rootDir = './ColorSets'
jsonFile = 'Contents.json'
# excelFileName = './colorSheet.xlsx'
excelFileName = './Darkmode_colorSheet_v0.1.xlsx'
excelSheetName = ''
deviceTypeStr = 'universal'
deviceTypes = [IdiomType.UNIVERSAL]
allSheets = OrderedDict()
allColorCount = 0
macCatalyst = False


def main():
    global rootDir
    global jsonFile
    global excelFileName
    global excelSheetName
    global deviceTypeStr
    global deviceTypes
    global macCatalyst

    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', metavar="EXCEL_FILE_NAME", dest="file", type=str, default=excelFileName, help=f"excel file name. default: {excelFileName}")
    parser.add_argument('-s', '--sheet', metavar="EXCEL_SHEET_NAME", dest="sheet", type=str, default=excelSheetName, help="excel sheet name. if dont't set name, all sheets colors will be extracted")
    parser.add_argument('-t', '--type', metavar="TYPE", dest="types", nargs='*', default=deviceTypeStr, help=f"target device types. default: {deviceTypeStr} / all types: [ universal, iphone, ipad, carplay, watch, tv, mac ] / ex) -t iphone ipad")
    parser.add_argument('-d', '--dir', metavar="DESTINATION_DIR", dest="directory", type=str, default=rootDir, help=f"target directory. default: {rootDir}")
    parser.add_argument('-c', '--catalyst', dest="catalyst", action='store_true', help="add colorset for mac-catalyst ex) -t ipad -c")

    args = parser.parse_args()
    excelFileName = args.file
    excelSheetName = args.sheet
    deviceTypeStr = args.types
    macCatalyst = args.catalyst
    rootDir = args.directory

    typeList = []
    deviceTypeList = deviceTypeStr.split()
    for deviceType in deviceTypeList:
        typeList.append(IdiomType[deviceType.upper().strip()])
    deviceTypes = typeList

    extract()


if __name__ == "__main__":
    main()
